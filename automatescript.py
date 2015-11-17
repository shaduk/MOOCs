

import openpyxl   #Import OpenPyxl module
wb = openpyxl.load_workbook('LOC54.xlsx')               #Open the excel file to be operated on
sheet = wb.get_sheet_by_name('Report')                  #Open the sub sheet "Report"
listpo = []             #Create a new list to store PO's
listps = []             #Create a new list to store Packing Slip's
listqty = []            #Create a new list to store quantity
i = 3                   #Scrap Sheet data from the third row
while(1==1):            #Keep going unless stopped by a break function
    i = i + 1           #Keep increasing the Rows by 1
    
    cellvalueqad = sheet['W'+str(i)].value    #Store the QAD PO value in a variable 
    cellvaluevss = sheet['A'+str(i)].value     #Store the VSS PO value in a variable
    if(sheet['R'+str(i)].value == None):      #If no value found in the R column of sheet break and stop the while loop
        break
    if(cellvalueqad !=cellvaluevss):        #Compare VSS PO and QAD PO in each line
        if(cellvaluevss != None):           # For all values in VSS PO which are not equal to QAD  
            print cellvaluevss
            print "Packing Slip: " + sheet['E'+str(i)].value
            print "Quantity :"+str(sheet['F'+str(i)].value)
            print "\n"
            listpo.append(str(cellvaluevss))        #Store QAD missing PO in listpo
            listps.append(str(sheet['E'+str(i)].value))         #Store Packing Slip of QAD missing PO
            listqty.append(str(sheet['F'+str(i)].value))        #Store Quantity of QAD missing PO
    if listpo:   #If there are any value in listpo 
        for j in range(0,len(listpo)):             #For every value in listpo if the po is found in the subsequent QAD columns with same qty remove it from listpo   
           
            if listpo[j]==str(cellvalueqad) and listqty[j]==str(sheet['AB'+str(i)].value):   #and listps[j]==str(sheet['AA'+str(i)].value)
                print listpo[j] + ' found in QAD with Packing Slip :' +  str(sheet['AA'+str(i)].value) + ' and Qty :' +str(sheet['AB'+str(i)].value)+'\n'
                listpo.remove(listpo[j])
                listps.remove(listps[j])
                listqty.remove(listqty[j])          
                break




for j in range(0,len(listpo)):
    print 'PO ' +listpo[j]+' with Packing Slip: '+listps[j]+' and qty: '+listqty[j]+' not found in qad' + '\n'

             
nb = openpyxl.load_workbook('recon.xlsx')       #open recon file to save data




def removedupli(seq):               #Remove duplicate POs from listpo 
    seen = set()                
    seen_add = seen.add
    return [ x for x in seq if not (x in seen or seen_add(x))]

povssdata = []          #Store VSS data in a list to write in excel
poqaddata = []          #Store QAD data in a list to write in excel
ponum = ''          
pops = ''               
poqty = 0
if listpo:                  #if any items in listpo
    listpo1 = removedupli(listpo)   #Remove duplicates from list PO and store in new variable
    for value in listpo1:
        i = 3                   #Start from 3rd row
        while(1==1):                #Keep continuing unless there is a break
            
            i = i + 1
            cellvalueqad = str(sheet['W'+str(i)].value)         #Store PO value in QAD in a variable
            cellvaluevss = str(sheet['A'+str(i)].value)          #Store PO value in QAD in a variable
            
            if(sheet['R'+str(i)].value == None):        #Stop when R column of excel sheet is none
                break
            if cellvaluevss==value:                     #For every PO in listpo1 get the data from corresponding columns in VSS and store in povssdata[] list
                
                pops = str(sheet['E'+str(i)].value)
                poqty = sheet['F'+str(i)].value
                povssdata.append([value,[pops,poqty]])
                pops = str(sheet['AA'+str(i)].value)
                poqty = sheet['AB'+str(i)].value
                poqaddata.append([value,[pops,poqty]])
            elif cellvalueqad==value:                   #For every PO in listpo1 get the data from corresponding columns in QAD and store in poqaddata[] list
                
                pops = str(sheet['E'+str(i)].value)
                poqty = sheet['F'+str(i)].value
                povssdata.append([value,[pops,poqty]])
                pops = str(sheet['AA'+str(i)].value)
                poqty = sheet['AB'+str(i)].value
                poqaddata.append([value,[pops,poqty]])


        


nb.create_sheet(index=0, title='LOC54')                 #Create a new sheet in recon file with the name
sheetnew = nb.get_sheet_by_name('LOC54')                #Select the newly created sheet

sheetnew['A1'] = 'PO Missing in QAD'
sheetnew['A1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['A'+str(i+2)] = povssdata[i][0]                #Save VSS PO data in the sheet
    

sheetnew['B1'] = 'Qty missing in QAD'
sheetnew['B1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    if(poqaddata[i][1][1] == None):
        sheetnew['B'+str(i+2)] = povssdata[i][1][1]                     
    else:
        sheetnew['B'+str(i+2)] = povssdata[i][1][1]-poqaddata[i][1][1]           #Save difference in qty data in the sheet
    
sheetnew['C1'] = 'Packaging Slip Number in VSS'
sheetnew['C1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['C'+str(i+2)] = povssdata[i][1][0]                      #Save VSS Packing Slip data in the sheet
    
sheetnew['D1'] = 'QTY In vSS'
sheetnew['D1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['D'+str(i+2)] = povssdata[i][1][1]              #Save VSS qty data in the sheet
    
sheetnew['E1'] = 'Packaging Slip Number in QAD'
sheetnew['E1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    sheetnew['E'+str(i+2)] = poqaddata[i][1][0]                 #Save QAD Packing Slip data in the sheet

sheetnew['F1'] = 'QTY in QAD'
sheetnew['F1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    sheetnew['F'+str(i+2)] = poqaddata[i][1][1]               #Save QAD qty data in the sheet
    
sheetnew['G1'] = 'Total in VSS'
sheetnew['G1'].font = openpyxl.styles.Font(size=12,bold=True)
sheetnew['H1'] = 'Total in QAD'
sheetnew['H1'].font = openpyxl.styles.Font(size=12,bold=True)

nb.save('recon.xlsx')
